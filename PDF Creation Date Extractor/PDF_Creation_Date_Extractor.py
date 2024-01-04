from PyPDF2 import PdfReader
import os
import openpyxl

EXCEL_NAME = 'Safety Observation Form'

def month_to_number(month_name):
    month_dict = {
        'Jan': 1,
        'Feb': 2,
        'Mar': 3,
        'Apr': 4,
        'May': 5,
        'Jun': 6,
        'Jul': 7,
        'Aug': 8,
        'Sep': 9,
        'Oct': 10,
        'Nov': 11,
        'Dec': 12
    }
    
    # Convert the month name to uppercase to handle case-insensitivity
    month_name = month_name.capitalize()

    # Check if the month name is in the dictionary
    if month_name in month_dict:
        return month_dict[month_name]
    else:
        return None  # Return None if the month name is not valid

def read_excel(write_vaule2, filename):
    day_month, year = write_vaule2.split(",", 2)[:2]
    year = year[1:]
    month_eng, day = day_month.split(" ")
    month = str(month_to_number(month_eng))
    write_vaule = "/".join([day, month, year])
    workbook = openpyxl.load_workbook(EXCEL_NAME+'.xlsx')

    sheet = workbook['Sheet1']

    # Read data row by row
    row_number = 2
    check = False
    for row in sheet.iter_rows(min_row=row_number, values_only=True):
        # print(f"Value: {row[0]} : {row[6]}")
        if row[0] == filename:
            # print(row[9])
            if row[9] is None:
                sheet.cell(row=row_number, column=10, value=write_vaule)
            check = True
        row_number += 1

    if not check:
        print(filename + " Not found!")
    workbook.save(EXCEL_NAME+'.xlsx')
    workbook.close()

current_directory = os.getcwd()

for foldername, subfolders, filenames in os.walk(current_directory):
    for filename in filenames:
        if filename.endswith('.pdf'):
            pdf_file_path = os.path.join(foldername, filename)
            print(f"Processing {filename}")
            getnextline = False
            stop = False
            with open(pdf_file_path, 'rb') as pdf_file:
                pdf_reader = PdfReader(pdf_file)

                # Iterate through pages
                for page in pdf_reader.pages:
                    page_text = page.extract_text()

                    # Split the page text into lines
                    lines = page_text.split('\n')
                    if stop:
                        break
                    # Process each line
                    for line in lines:
                        if "created on:" in line.lower():
                            x = line.split(":")[-1]
                        if getnextline:
                            #print(line.replace(x, ""))
                            read_excel(line.replace(x, ""), filename.split(".")[0])
                            getnextline = False
                            stop = True
                        if "history" in line.lower():
                            getnextline = True


